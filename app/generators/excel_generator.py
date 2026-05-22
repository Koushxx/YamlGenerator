"""
Excel template generator and parser module.
Handles creation of device-specific Excel templates and parsing uploaded Excel files.
"""

import os
from typing import Dict, List, Any, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.templates.base_template import BaseTemplate


# Styling constants
HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
MANDATORY_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
DESC_FONT = Font(name="Segoe UI", italic=True, size=10, color="555555")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_excel_template(template: BaseTemplate, output_path: str) -> str:
    """
    Generate an Excel template file for the given device template.
    Creates sheets: Device Configuration, Tags, Bundling.
    Returns the path to the created file.
    """
    wb = Workbook()

    # --- Device Configuration Sheet ---
    ws_device = wb.active
    ws_device.title = "Device Configuration"
    _write_sheet_headers(ws_device, template.device_columns)
    _add_sample_row(ws_device, template.device_columns, row=3)

    # --- Tags Sheet ---
    ws_tags = wb.create_sheet("Tags")
    _write_sheet_headers(ws_tags, template.tag_columns)
    _add_sample_row(ws_tags, template.tag_columns, row=3)

    # --- Bundling Sheet ---
    ws_bundling = wb.create_sheet("Bundling")
    _write_sheet_headers(ws_bundling, template.bundling_columns)
    _add_sample_row(ws_bundling, template.bundling_columns, row=3)

    # --- Instructions Sheet ---
    ws_instructions = wb.create_sheet("Instructions")
    _write_instructions(ws_instructions, template)

    wb.save(output_path)
    return output_path


def _write_sheet_headers(ws, columns: List[Dict[str, Any]]):
    """Write column headers and description row."""
    for col_idx, col_def in enumerate(columns, start=1):
        # Header row
        cell = ws.cell(row=1, column=col_idx, value=col_def["name"])
        cell.font = HEADER_FONT
        cell.fill = MANDATORY_FILL if col_def["mandatory"] else OPTIONAL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

        # Description row
        mandatory_mark = " *" if col_def["mandatory"] else ""
        desc_cell = ws.cell(row=2, column=col_idx, value=f"{col_def['description']}{mandatory_mark}")
        desc_cell.font = DESC_FONT
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        desc_cell.border = THIN_BORDER

        # Auto-width
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(20, len(col_def["name"]) + 8)

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 40


def _add_sample_row(ws, columns: List[Dict[str, Any]], row: int):
    """Add a sample/default row to guide users."""
    for col_idx, col_def in enumerate(columns, start=1):
        default_val = col_def.get("default")
        if default_val is not None:
            ws.cell(row=row, column=col_idx, value=default_val)


def _write_instructions(ws, template: BaseTemplate):
    """Write instructions sheet with usage guidance."""
    ws.column_dimensions["A"].width = 80
    instructions = [
        f"Device YAML Generator - {template.display_name} Template",
        "",
        "INSTRUCTIONS:",
        "─" * 60,
        "1. Fill in the 'Device Configuration' sheet with your device connection parameters.",
        "   - Fields marked with (*) in the description row are MANDATORY.",
        "   - Row 2 contains descriptions - DO NOT edit row 1 or row 2.",
        "   - Enter your data starting from row 3.",
        "",
        "2. Fill in the 'Tags' sheet with all tags/variables you want to monitor.",
        "   - Each row represents one tag.",
        "   - Enter data starting from row 3.",
        "",
        "3. (Optional) Fill in the 'Bundling' sheet to group tags for transmission.",
        "   - Mode can be: cyclic (time-based) or trigger (event-based).",
        "   - For cyclic mode: set 'Cyclic (ms)' column.",
        "   - For trigger mode: set 'Trigger Tag' and optionally 'Trigger Mode'.",
        "   - Data Tags: comma-separated tag names to include in the bundle.",
        "",
        "4. Save the file and upload it back to the Device YAML Generator application.",
        "",
        "NOTES:",
        "─" * 60,
        f"• Device Type: {template.device_type}",
        "• Do NOT rename or reorder the sheet tabs.",
        "• Do NOT modify header row (row 1).",
        "• Leave optional fields empty if not needed.",
    ]

    for row_idx, line in enumerate(instructions, start=1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = Font(name="Segoe UI", bold=True, size=14)
        elif line.startswith("─"):
            cell.font = Font(name="Segoe UI", size=10, color="888888")
        else:
            cell.font = Font(name="Segoe UI", size=11)


def parse_excel_file(file_path: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Parse an uploaded Excel file and return device data, tags list, and bundling list.
    Returns: (device_data_dict, tags_list, bundling_list)
    Raises ValueError for invalid formats.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path, data_only=True)

    # Parse Device Configuration sheet
    if "Device Configuration" not in wb.sheetnames:
        raise ValueError("Missing required sheet: 'Device Configuration'")
    device_data = _parse_sheet_single_row(wb["Device Configuration"])

    # Parse Tags sheet
    if "Tags" not in wb.sheetnames:
        raise ValueError("Missing required sheet: 'Tags'")
    tags_data = _parse_sheet_multiple_rows(wb["Tags"])

    # Parse Bundling sheet (optional)
    bundling_data = []
    if "Bundling" in wb.sheetnames:
        bundling_data = _parse_sheet_multiple_rows(wb["Bundling"])

    wb.close()
    return device_data, tags_data, bundling_data


def _parse_sheet_single_row(ws) -> Dict[str, Any]:
    """Parse a sheet where row 1 = headers, row 2 = descriptions, row 3+ = data (take first data row)."""
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers.append(str(val).strip())
        else:
            break

    if not headers:
        raise ValueError(f"No headers found in sheet '{ws.title}'")

    data = {}
    for col_idx, header in enumerate(headers, start=1):
        val = ws.cell(row=3, column=col_idx).value
        if val is not None:
            data[header] = val

    return data


def _parse_sheet_multiple_rows(ws) -> List[Dict[str, Any]]:
    """Parse a sheet where row 1 = headers, row 2 = descriptions, row 3+ = data rows."""
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers.append(str(val).strip())
        else:
            break

    if not headers:
        return []

    rows = []
    for row in range(3, ws.max_row + 1):
        row_data = {}
        has_data = False
        for col_idx, header in enumerate(headers, start=1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                row_data[header] = val
                has_data = True
        if has_data:
            rows.append(row_data)

    return rows
